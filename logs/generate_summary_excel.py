"""Generate summary Excel with eval mode and protected mode results."""
import json
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

LOGS_DIR = os.path.dirname(os.path.abspath(__file__))
SEEDS = [42, 43, 44, 45, 46]
AGENTS = [100, 200, 500]

EVAL_DIR = LOGS_DIR  # eval mode results (protection=0)
PROT_DIR = os.path.join(LOGS_DIR, "protected")  # protected results (protection=24)


def load_summary(base_dir, seed, n):
    path = os.path.join(base_dir, f"seed_{seed}", f"N{n}", "summary.json")
    if not os.path.exists(path):
        return None
    with open(path, "r") as f:
        return json.load(f)


def avg(values):
    return sum(values) / len(values) if values else 0.0


def std(values):
    if len(values) < 2:
        return 0.0
    m = avg(values)
    return (sum((v - m) ** 2 for v in values) / (len(values) - 1)) ** 0.5


def collect_metrics(base_dir):
    results = {}
    for n in AGENTS:
        metrics_per_seed = []
        for seed in SEEDS:
            s = load_summary(base_dir, seed, n)
            if s is None:
                continue
            metrics_per_seed.append({
                "precision": s.get("precision", 0),
                "recall": s.get("recall", 0),
                "f1": s.get("f1", 0),
                "accuracy": s.get("accuracy", 0),
                "tpr": s.get("tpr", 0),
                "tnr": s.get("tnr", 0),
                "fpr": s.get("fpr", 0),
                "fnr": s.get("fnr", 0),
                "attack_rate_reduction": s.get("attack_rate_reduction", 0),
                "cost_efficiency": s.get("cost_efficiency", 0),
                "dynamic_attack_rate": s.get("dynamic_mean_attack_rate", 0),
                "baseline_attack_rate": s.get("baseline_mean_attack_rate", 0),
                "dynamic_audit_cost": s.get("dynamic_total_audit_cost", 0),
                "baseline_audit_cost": s.get("baseline_total_audit_cost", 0),
                "risk_mitigation": s.get("risk_mitigation_truth_based", 0),
                "tp": s.get("tp", 0),
                "tn": s.get("tn", 0),
                "fp": s.get("fp", 0),
                "fn": s.get("fn", 0),
            })
            # Per-attack metrics
            pa = s.get("per_attack_metrics", {})
            for atype in ["FDI", "DOS", "MITM", "FAULT", "CHAIN"]:
                ad = pa.get(atype, {})
                metrics_per_seed[-1][f"{atype}_tpr"] = ad.get("tpr", 0)
                metrics_per_seed[-1][f"{atype}_support"] = ad.get("support", 0)

        if not metrics_per_seed:
            continue

        agg = {}
        for key in metrics_per_seed[0]:
            vals = [m[key] for m in metrics_per_seed]
            agg[f"{key}_mean"] = avg(vals)
            agg[f"{key}_std"] = std(vals)
        agg["num_seeds"] = len(metrics_per_seed)
        results[n] = agg
    return results


def write_excel(eval_results, prot_results, output_path):
    if not HAS_OPENPYXL:
        print("openpyxl not available, writing CSV instead")
        write_csv(eval_results, prot_results, output_path.replace(".xlsx", ".csv"))
        return

    wb = Workbook()

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    prot_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    number_fmt_pct = "0.00%"
    number_fmt_dec = "0.0000"
    number_fmt_int = "#,##0"
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # --- Sheet 1: Binary Detection Metrics ---
    ws = wb.active
    ws.title = "Binary Detection"

    headers = ["Metric", "N=100 (Eval)", "N=200 (Eval)", "N=500 (Eval)",
               "N=100 (Protected)", "N=200 (Protected)", "N=500 (Protected)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    metrics_rows = [
        ("Recall (TPR)", "recall", number_fmt_pct),
        ("Precision", "precision", number_fmt_pct),
        ("F1-Score", "f1", number_fmt_pct),
        ("Accuracy", "accuracy", number_fmt_pct),
        ("FPR", "fpr", number_fmt_pct),
        ("TNR", "tnr", number_fmt_pct),
        ("FNR", "fnr", number_fmt_pct),
        ("True Positives", "tp", number_fmt_int),
        ("True Negatives", "tn", number_fmt_int),
        ("False Positives", "fp", number_fmt_int),
        ("False Negatives", "fn", number_fmt_int),
    ]

    for row_idx, (label, key, fmt) in enumerate(metrics_rows, 2):
        ws.cell(row=row_idx, column=1, value=label).border = thin_border
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        for col_offset, n in enumerate(AGENTS):
            # Eval mode
            col = col_offset + 2
            if n in eval_results:
                mean_val = eval_results[n].get(f"{key}_mean", 0)
                std_val = eval_results[n].get(f"{key}_std", 0)
                c = ws.cell(row=row_idx, column=col)
                c.value = mean_val
                c.number_format = fmt
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")
            # Protected mode
            col = col_offset + 5
            if prot_results and n in prot_results:
                mean_val = prot_results[n].get(f"{key}_mean", 0)
                c = ws.cell(row=row_idx, column=col)
                c.value = mean_val
                c.number_format = fmt
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")
                c.fill = prot_fill

    ws.column_dimensions["A"].width = 20
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # --- Sheet 2: Cost & Attack Reduction ---
    ws2 = wb.create_sheet("Cost Efficiency")

    headers2 = ["Metric", "N=100 (Eval)", "N=200 (Eval)", "N=500 (Eval)",
                "N=100 (Protected)", "N=200 (Protected)", "N=500 (Protected)"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    cost_rows = [
        ("Attack Rate (Dynamic)", "dynamic_attack_rate", number_fmt_pct),
        ("Attack Rate (Baseline)", "baseline_attack_rate", number_fmt_pct),
        ("Attack Rate Reduction", "attack_rate_reduction", number_fmt_pct),
        ("Dynamic Audit Cost", "dynamic_audit_cost", number_fmt_int),
        ("Baseline Audit Cost", "baseline_audit_cost", number_fmt_int),
        ("Cost Efficiency (savings)", "cost_efficiency", number_fmt_pct),
        ("Risk Mitigation", "risk_mitigation", number_fmt_pct),
    ]

    for row_idx, (label, key, fmt) in enumerate(cost_rows, 2):
        ws2.cell(row=row_idx, column=1, value=label).border = thin_border
        ws2.cell(row=row_idx, column=1).font = Font(bold=True)
        for col_offset, n in enumerate(AGENTS):
            col = col_offset + 2
            if n in eval_results:
                c = ws2.cell(row=row_idx, column=col)
                c.value = eval_results[n].get(f"{key}_mean", 0)
                c.number_format = fmt
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")
            col = col_offset + 5
            if prot_results and n in prot_results:
                c = ws2.cell(row=row_idx, column=col)
                c.value = prot_results[n].get(f"{key}_mean", 0)
                c.number_format = fmt
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")
                c.fill = prot_fill

    ws2.column_dimensions["A"].width = 28
    for col in range(2, 8):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # --- Sheet 3: Per-Attack TPR ---
    ws3 = wb.create_sheet("Per-Attack TPR")

    headers3 = ["Attack Type", "N=100 (Eval)", "N=200 (Eval)", "N=500 (Eval)",
                "N=100 (Protected)", "N=200 (Protected)", "N=500 (Protected)"]
    for col, h in enumerate(headers3, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    attack_types = ["FDI", "DOS", "MITM", "FAULT", "CHAIN"]
    for row_idx, atype in enumerate(attack_types, 2):
        ws3.cell(row=row_idx, column=1, value=atype).border = thin_border
        ws3.cell(row=row_idx, column=1).font = Font(bold=True)
        for col_offset, n in enumerate(AGENTS):
            col = col_offset + 2
            if n in eval_results:
                tpr = eval_results[n].get(f"{atype}_tpr_mean", 0)
                support = eval_results[n].get(f"{atype}_support_mean", 0)
                c = ws3.cell(row=row_idx, column=col)
                c.value = tpr
                c.number_format = number_fmt_pct
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")
            col = col_offset + 5
            if prot_results and n in prot_results:
                tpr = prot_results[n].get(f"{atype}_tpr_mean", 0)
                c = ws3.cell(row=row_idx, column=col)
                c.value = tpr
                c.number_format = number_fmt_pct
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")
                c.fill = prot_fill

    # Support counts row
    row_idx = len(attack_types) + 3
    ws3.cell(row=row_idx, column=1, value="Ground Truth Counts (avg)").font = Font(bold=True, italic=True)
    row_idx += 1
    for ridx, atype in enumerate(attack_types, row_idx):
        ws3.cell(row=ridx, column=1, value=f"{atype} support").border = thin_border
        for col_offset, n in enumerate(AGENTS):
            col = col_offset + 2
            if n in eval_results:
                c = ws3.cell(row=ridx, column=col)
                c.value = eval_results[n].get(f"{atype}_support_mean", 0)
                c.number_format = number_fmt_int
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")

    ws3.column_dimensions["A"].width = 24
    for col in range(2, 8):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    # --- Sheet 4: Cross-Seed Variance ---
    ws4 = wb.create_sheet("Cross-Seed Variance")

    headers4 = ["Metric", "N=100 Mean +/- Std", "N=200 Mean +/- Std", "N=500 Mean +/- Std"]
    for col, h in enumerate(headers4, 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    var_metrics = [
        ("Recall", "recall"),
        ("Precision", "precision"),
        ("F1", "f1"),
        ("Accuracy", "accuracy"),
        ("FPR", "fpr"),
        ("Attack Rate Reduction", "attack_rate_reduction"),
        ("Cost Efficiency", "cost_efficiency"),
    ]

    for row_idx, (label, key) in enumerate(var_metrics, 2):
        ws4.cell(row=row_idx, column=1, value=label).border = thin_border
        ws4.cell(row=row_idx, column=1).font = Font(bold=True)
        for col_offset, n in enumerate(AGENTS):
            col = col_offset + 2
            if n in eval_results:
                m = eval_results[n].get(f"{key}_mean", 0)
                s = eval_results[n].get(f"{key}_std", 0)
                c = ws4.cell(row=row_idx, column=col)
                c.value = f"{m*100:.2f}% +/- {s*100:.2f}%"
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")

    ws4.column_dimensions["A"].width = 24
    for col in range(2, 5):
        ws4.column_dimensions[get_column_letter(col)].width = 24

    wb.save(output_path)
    print(f"Saved: {output_path}")


def write_csv(eval_results, prot_results, output_path):
    import csv
    with open(output_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Mode", "N", "Recall", "Precision", "F1", "Accuracy", "FPR",
                     "Attack_Rate_Reduction", "Cost_Efficiency", "Risk_Mitigation",
                     "FDI_TPR", "DOS_TPR", "MITM_TPR", "FAULT_TPR",
                     "Dynamic_Audit_Cost", "Baseline_Audit_Cost", "Seeds"])
        for mode, results in [("Eval (no protection)", eval_results), ("Protected (window=24)", prot_results)]:
            if not results:
                continue
            for n in AGENTS:
                if n not in results:
                    continue
                r = results[n]
                w.writerow([
                    mode, n,
                    f"{r.get('recall_mean', 0):.4f}",
                    f"{r.get('precision_mean', 0):.4f}",
                    f"{r.get('f1_mean', 0):.4f}",
                    f"{r.get('accuracy_mean', 0):.4f}",
                    f"{r.get('fpr_mean', 0):.4f}",
                    f"{r.get('attack_rate_reduction_mean', 0):.4f}",
                    f"{r.get('cost_efficiency_mean', 0):.4f}",
                    f"{r.get('risk_mitigation_mean', 0):.4f}",
                    f"{r.get('FDI_tpr_mean', 0):.4f}",
                    f"{r.get('DOS_tpr_mean', 0):.4f}",
                    f"{r.get('MITM_tpr_mean', 0):.4f}",
                    f"{r.get('FAULT_tpr_mean', 0):.4f}",
                    f"{r.get('dynamic_audit_cost_mean', 0):.0f}",
                    f"{r.get('baseline_audit_cost_mean', 0):.0f}",
                    r.get("num_seeds", 0),
                ])
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    eval_results = collect_metrics(EVAL_DIR)
    prot_results = collect_metrics(PROT_DIR) if os.path.exists(PROT_DIR) else {}

    output = os.path.join(LOGS_DIR, "experiment_summary.xlsx")
    if HAS_OPENPYXL:
        write_excel(eval_results, prot_results, output)
    else:
        output = os.path.join(LOGS_DIR, "experiment_summary.csv")
        write_csv(eval_results, prot_results, output)
